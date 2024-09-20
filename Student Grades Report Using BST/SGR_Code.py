import openpyxl  # Library for working with Excel files
from openpyxl import Workbook  # To create a new workbook
import os  # Library to check if the Excel file exists
import pandas as pd  # Library for working with DataFrames

# Class to represent the elements of the binary search tree (student code and record number)
class BinTreeElementType:
    def __init__(self, code, recNo):
        self.code = code  # Unique code for each student
        self.recNo = recNo  # Record number (used for ordering in BST)

# Class to represent a node in the binary search tree
class BinTreeNode:
    def __init__(self, data):
        self.data = data  # Data contains student code and record number
        self.LChild = None  # Left child of the node
        self.RChild = None  # Right child of the node

# Class to represent a student with various attributes
class StudentT:
    def __init__(self, code, surname, name, sex, year, grade):
        self.code = code  # Student's unique code
        self.surname = surname  # Student's surname
        self.name = name  # Student's name
        self.sex = sex  # Student's gender (F/M)
        self.year = year  # Student's registration year
        self.grade = grade  # Student's grade (0-20)

# Function to create a new binary search tree (BST) (initially empty)
def CreateBST():
    return None

# Function to check if the BST is empty
def BSTEmpty(Root):
    return Root is None  # Return True if the tree is empty, otherwise False

# Function to insert an item into the BST recursively
def RecBSTInsert(Root, Item):
    if BSTEmpty(Root):
        return BinTreeNode(Item)
    else:
        if Item.code < Root.data.code:
            Root.LChild = RecBSTInsert(Root.LChild, Item)
        elif Item.code > Root.data.code:
            Root.RChild = RecBSTInsert(Root.RChild, Item)
        return Root

# Function to search for a specific item in the BST recursively
def RecBSTSearch(Root, KeyValue):
    if BSTEmpty(Root):
        return False, None
    elif KeyValue.code < Root.data.code:
        return RecBSTSearch(Root.LChild, KeyValue)
    elif KeyValue.code > Root.data.code:
        return RecBSTSearch(Root.RChild, KeyValue)
    else:
        return True, Root

# Function to display the menu and take user input for the choice
def menu():
    print("\n                  MENU                  \n")
    print("-------------------------------------------------\n")
    print("1. Insert new student\n")
    print("2. Search for a student\n")
    print("3. Print all students (Traverse Inorder)\n")
    print("4. Get records within a specific grade range\n")
    print("5. Quit\n")
    choice = int(input("\nChoice: "))
    return choice

# Function to load or create the Excel file for storing student data
def load_excel_data(filename):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["Code", "Surname", "Name", "Sex", "Year", "Grade"])  # Add headers to the sheet
        wb.save(filename)
        return wb, ws
    else:
        wb = openpyxl.load_workbook(filename)
        ws = wb["Students"]
        return wb, ws

# Function to load student records from the Excel file into a list
def load_students_from_excel(filename):
    wb, ws = load_excel_data(filename)
    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        code, surname, name, sex, year, grade = row
        students.append(StudentT(code, surname, name, sex, year, grade))
    return students

# Function to save a new student record into the Excel file in sorted order
def save_student_to_excel(filename, student):
    students = load_students_from_excel(filename)
    students.append(student)
    students.sort(key=lambda s: s.code)

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Code", "Surname", "Name", "Sex", "Year", "Grade"])

    for s in students:
        ws.append([s.code, s.surname, s.name, s.sex, s.year, s.grade])

    wb.save(filename)
    print("Student saved successfully to Excel file in sorted order.")

# Function to print all students using DataFrame
def print_all_students_as_dataframe(filename):
    students = load_students_from_excel(filename)
    if not students:
        print("No one enrolled.")
        return
    df = pd.DataFrame([vars(s) for s in students])
    print(df)

# Function to print students within a specific grade range using DataFrame
def print_students_within_grade_range_as_dataframe(filename, low, high):
    students = load_students_from_excel(filename)
    filtered_students = [s for s in students if low <= s.grade <= high]
    if not filtered_students:
        print(f"No students found with grades between {low} and {high}.")
        return
    df = pd.DataFrame([vars(s) for s in filtered_students])
    print(df)

# Main program function
def main():
    Root = CreateBST()  # Create an empty binary search tree
    filename = "Students_Report.xlsx"  # Excel file to store student records
    
    while True:
        choice = menu()  # Show menu and get the user's choice
        
        if choice == 1:  # Insert a new student record
            code = int(input("Give student's AM: "))
            while code < 0:  # Ensure student code is non-negative
                print("Code can't be a negative number")
                code = int(input("Give student's AM: "))

            found, _ = RecBSTSearch(Root, BinTreeElementType(code, 0))
            if found:  # If student code is found, prevent insertion
                print("This code is already assigned to another student.")
            else:
                surname = input("Give student surname: ").strip()
                name = input("Give student name: ").strip()
                year = int(input("Give student's registration year: "))
                
                grade = float(input("Give student's grade (0-20): "))
                while grade < 0 or grade > 20:
                    print("Grade must be between 0 and 20.")
                    grade = float(input("Give student's grade (0-20): "))
                
                sex = input("Give student sex (F/M): ").upper().strip()
                while sex not in ['F', 'M']:
                    print("Sex must be F or M.")
                    sex = input("Give student sex (F/M): ").upper().strip()

                student = StudentT(code, surname, name, sex, year, grade)
                save_student_to_excel(filename, student)  # Save the student record to Excel

                indexRec = BinTreeElementType(code, len(load_students_from_excel(filename)))
                Root = RecBSTInsert(Root, indexRec)

        elif choice == 2:  # Search for a student by code
            code = int(input("Give student's code: "))
            found, LocPtr = RecBSTSearch(Root, BinTreeElementType(code, 0))
            if found:
                student = next(s for s in load_students_from_excel(filename) if s.code == code)
                print(f"Student found: Code: {student.code}, Surname: {student.surname}, Name: {student.name}, "
                      f"Sex: {student.sex}, Year: {student.year}, Grade: {student.grade}")
            else:
                print("No such records are present.")

        elif choice == 3:  # Print all students
            print_all_students_as_dataframe(filename)

        elif choice == 4:  # Get records within a specific grade range
            low = float(input("Enter the lower bound of the grade range: "))
            high = float(input("Enter the upper bound of the grade range: "))
            print_students_within_grade_range_as_dataframe(filename, low, high)

        elif choice == 5:  # Exit the program
            print("Exiting the program.")
            break

        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()  # Start the program
